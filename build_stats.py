"""Build stats_latest.json for the CARTA Our Impact dashboard.

Reads the daily-refreshed `_latest.xlsx` files in ./data/ and emits
`data/stats_latest.json` containing:

- `rows`: cleaned row-level data for fellows, postdocs, grants, and trainings
  (the page re-aggregates client-side for Power-BI-style cross-filtering).
- `measures_static`: numbers that are not derivable from `rows` (currently empty —
  every published measure is derivable).
- `meta`: source filenames, generated_at, and counts for quick sanity checks.

Run locally:  python build_stats.py
In CI:        invoked from .github/workflows/download-excel.yml after the
              download step succeeds.
"""
import datetime as dt
import json
import re
import sys
import warnings
from pathlib import Path

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning)

DATA_DIR = Path(__file__).resolve().parent / "data"
OUT_FILE = DATA_DIR / "stats_latest.json"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def read_sheet(xlsx_path: Path, sheet_name: str) -> list[dict]:
    """Return a list of dicts from the named sheet, keyed by trimmed header."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise KeyError(f"{xlsx_path.name}: sheet '{sheet_name}' missing. Have: {wb.sheetnames}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(rows)]
    out = []
    for row in rows:
        if all(v is None for v in row):
            continue
        out.append({h: v for h, v in zip(header, row) if h})
    wb.close()
    return out


def to_int(v, default=None):
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(str(v).strip())
    except (ValueError, TypeError):
        return default


def to_float(v, default=None):
    if v is None or v == "":
        return default
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return default


def to_year(v):
    """Coerce a value to a 4-digit year int, or None."""
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.year
    if isinstance(v, dt.date):
        return v.year
    if isinstance(v, (int, float)):
        n = int(v)
        return n if 1990 <= n <= 2100 else None
    s = str(v).strip()
    m = re.search(r"(19|20)\d{2}", s)
    return int(m.group(0)) if m else None


def normalize_status(v: str | None) -> str | None:
    """Map various status strings to a canonical set: Completed / In progress / Terminated."""
    if not v:
        return None
    s = str(v).strip().lower()
    if "complete" in s or "defend" in s or "graduat" in s:
        return "Completed"
    if "progress" in s:
        return "In progress"
    if "terminat" in s:
        return "Terminated"
    return str(v).strip()


def normalize_gender(v: str | None) -> str | None:
    if not v:
        return None
    s = str(v).strip().lower()
    if s.startswith("f"):
        return "Female"
    if s.startswith("m"):
        return "Male"
    return None


def normalize_intervention(v: str | None) -> str | None:
    if not v:
        return None
    s = str(v).strip().upper()
    for tag in ("JAS", "SW", "APAS", "GGWW"):
        if s == tag or s.startswith(tag + " ") or s.startswith(tag + "-"):
            return tag
    if "PHD" in s:
        return "PhD training"
    return str(v).strip()


def normalize_training_type(v: str | None) -> str | None:
    """Map type-of-training labels to canonical values."""
    if not v:
        return None
    s = str(v).strip().lower()
    if "joint" in s:
        return "Joint training"
    if "common tot" in s or "common to t" in s:
        return "Common ToT"
    if "local tot" in s or "local to t" in s:
        return "Local ToT"
    if "local training" in s:
        return "Local training"
    if "general workshop" in s:
        return "General workshop"
    return str(v).strip()


def normalize_duty(v: str | None) -> str:
    """ToT facilitator vs participant. Reduce composite labels to primary role."""
    if not v:
        return "Participant"
    s = str(v).strip().lower()
    if "facilit" in s or "organizer" in s or "management" in s:
        return "Facilitator"
    return "Participant"


# ---------------------------------------------------------------------------
# Per-source extractors
# ---------------------------------------------------------------------------


def load_fellows() -> list[dict]:
    """Use the cohort 1-11 demographics file (most complete)."""
    src = DATA_DIR / "Cohort_1_11_Demographics_latest.xlsx"
    rows = read_sheet(src, "Fellows")
    out = []
    for r in rows:
        unique_id = r.get("Unique ID")
        if not unique_id:
            continue
        out.append({
            "id": str(unique_id),
            "gender": normalize_gender(r.get("Gender")),
            "cohort": to_int(r.get("Cohort")),
            "nationality": (r.get("Nationality") or "").strip() or None,
            "institution_employment": (r.get("Institution of employment at registration") or "").strip() or None,
            "institution_registration": (r.get("Institution of registration") or "").strip() or None,
            "year_admission": to_year(r.get("Year of admission into CARTA")) or to_year(r.get("Date of PhD registration")),
            "year_completion": to_year(r.get("Date of completion(Defended/Graduated)")),
            "status": normalize_status(r.get("Current PhD Status ( Completed/Defended/In Progress)")),
            "ttc_months": to_float(r.get("Time to completion since PhD registration (Months)")),
            "promotion": (r.get("Promotion event") or "").strip() or None,
            "responsibilities": (str(r.get("Other responsibilities") or "").strip() or None),
            "pubs_during_phd": to_int(r.get("No of Publications During PhD")),
            "pubs_after_phd": to_int(r.get("No of Publications after PhD")),
            "first_author_phd_pubs": to_int(r.get("1st Author PhD  Publications")),
            "last_author_phd_pubs": to_int(r.get("Last Author PhD Publications")),
            "first_author_grad_pubs": to_int(r.get("1st Author Graduate  Publications")),
            "last_author_grad_pubs": to_int(r.get("Last Author  Graduate Publications")),
            "terminated_date": to_year(r.get("Terminated")),
            "funder": (str(r.get("Fellow Funder") or "").strip() or None),
            "jas_attended": sum(
                1 for k in ("Month/ Year JAS1", "Month/ Year JAS2", "Month/ Year JAS3", "Month/ Year JAS4")
                if r.get(k) not in (None, "")
            ),
        })
    return out


def load_postdocs() -> list[dict]:
    rows = read_sheet(DATA_DIR / "Postdocs_latest.xlsx", "Post Doc")
    out = []
    for r in rows:
        if not r.get("Unique ID") and not r.get("Name of Awardee"):
            continue
        out.append({
            "id": str(r.get("Unique ID") or ""),
            "sex": normalize_gender(r.get("Sex")),
            "institution_employment": (r.get("Institution of employment at the time of award") or "").strip() or None,
            "host_country": (r.get("Host Country") or "").strip() or None,
            "award_type": (r.get("Award Type") or "").strip() or None,
            "year_award": to_year(r.get("Year of Award")),
            "year_completion": to_year(r.get("Year of Completion")),
            "status": normalize_status(r.get("Status (Active/Completed")),
            "funder": (r.get("Funder") or "").strip() or None,
        })
    return out


def load_grants() -> list[dict]:
    rows = read_sheet(DATA_DIR / "Extra Grants_latest.xlsx", "Extra Grants")
    out = []
    for r in rows:
        amount = to_float(r.get("Total amount in $"))
        if amount is None:
            # Power BI sums by amount; drop rows we can't sum. Counts already preserved upstream.
            continue
        out.append({
            "sex": normalize_gender(r.get("Sex")),
            "cohort": to_int(r.get("Cohort Number")),
            "institution": (r.get("Institution of employment at registration") or "").strip() or None,
            "type": (r.get("Type of Grant") or "").strip() or None,
            "year": to_year(r.get("Year")) or to_year(r.get("Award date")),
            "amount_usd": amount,
            "funder": (r.get("Name of Funder") or "").strip() or None,
            "duration_months": to_int(r.get("Duration of Funding (in months)")),
        })
    return out


def load_trainings():
    src = DATA_DIR / "Institutionalization_latest.xlsx"

    def _load(sheet_name, source_tag):
        rows = read_sheet(src, sheet_name)
        out = []
        for r in rows:
            if not r.get("Full Name") and not r.get("Intervention"):
                continue
            out.append({
                "source": source_tag,  # "carta" | "institutional"
                "intervention": normalize_intervention(r.get("Intervention")),
                "type": normalize_training_type(r.get("Type of training")),
                "duty": normalize_duty(r.get("Principal duty at event (Participant/Facilitator)")),
                "sex": normalize_gender(r.get("Sex")),
                "year": to_year(r.get("Year")) or to_year(r.get("Event Start Date")),
                "institution": (r.get("Associated institution") or "").strip() or None,
                "is_carta_fellow": (str(r.get("CARTA Fellow (Yes/No)") or "").strip().lower().startswith("y")),
            })
        return out

    return _load("CARTA Organized", "carta") + _load("Local and Institutional ToTs", "institutional")


def load_curricula_institutions(trainings: list[dict]) -> list[str]:
    """Derive list of institutions that have adopted CARTA curricula from institutional trainings."""
    from collections import Counter
    counts = Counter()
    for t in trainings:
        if t["source"] == "institutional" and t["institution"]:
            counts[t["institution"]] += 1
    # Threshold: at least 5 training-events at that institution to count as "adopted"
    return sorted(name for name, n in counts.items() if n >= 5)


# ---------------------------------------------------------------------------
# Aggregate sanity-check measures (for meta block only — page recomputes everything)
# ---------------------------------------------------------------------------


def quick_measures(fellows, postdocs, grants, trainings) -> dict:
    total_fellows = len(fellows)
    by_status = {}
    for f in fellows:
        by_status[f["status"] or "Unknown"] = by_status.get(f["status"] or "Unknown", 0) + 1
    by_gender = {}
    for f in fellows:
        by_gender[f["gender"] or "Unknown"] = by_gender.get(f["gender"] or "Unknown", 0) + 1
    ttcs = [f["ttc_months"] for f in fellows if f["ttc_months"]]
    avg_ttc = round(sum(ttcs) / len(ttcs), 1) if ttcs else None
    median_ttc = None
    if ttcs:
        s = sorted(ttcs)
        median_ttc = round(s[len(s) // 2] if len(s) % 2 else (s[len(s)//2 - 1] + s[len(s)//2]) / 2, 1)
    total_grants = round(sum(g["amount_usd"] for g in grants))
    pubs = sum((f["pubs_during_phd"] or 0) + (f["pubs_after_phd"] or 0) for f in fellows)
    completed_fellows = by_status.get("Completed", 0)
    retention_rate = round(((total_fellows - by_status.get("Terminated", 0)) / total_fellows) * 100, 1) if total_fellows else None

    # Training counts (by intervention × type × source)
    from collections import defaultdict
    training_counts = defaultdict(int)
    for t in trainings:
        if not t["intervention"] or not t["type"]:
            continue
        key = f"{t['source']}::{t['intervention']}::{t['type']}"
        training_counts[key] += 1
    jas_person_events = sum(f.get("jas_attended", 0) or 0 for f in fellows)

    return {
        "total_fellows": total_fellows,
        "completed": completed_fellows,
        "in_progress": by_status.get("In progress", 0),
        "terminated": by_status.get("Terminated", 0),
        "fellows_by_gender": by_gender,
        "avg_ttc_months": avg_ttc,
        "median_ttc_months": median_ttc,
        "retention_rate_pct": retention_rate,
        "total_postdocs": len(postdocs),
        "postdocs_completed": sum(1 for p in postdocs if p["status"] == "Completed"),
        "extra_grants_usd": total_grants,
        "extra_grants_count": len(grants),
        "peer_reviewed_articles": pubs,
        "jas_person_events": jas_person_events,
        "training_counts": dict(training_counts),
    }


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main():
    print("Building stats_latest.json...")
    fellows = load_fellows()
    print(f"  Fellows: {len(fellows)}")
    postdocs = load_postdocs()
    print(f"  Postdocs: {len(postdocs)}")
    grants = load_grants()
    print(f"  Grants (with amounts): {len(grants)}")
    trainings = load_trainings()
    print(f"  Training records: {len(trainings)}")
    institutions = load_curricula_institutions(trainings)
    print(f"  Curricula institutions: {len(institutions)}")

    measures = quick_measures(fellows, postdocs, grants, trainings)
    print(f"  Quick measures: total_fellows={measures['total_fellows']}, "
          f"completed={measures['completed']}, grants=${measures['extra_grants_usd']:,}")

    payload = {
        "generated_at": dt.datetime.now(dt.UTC).isoformat(timespec="seconds"),
        "source_files": {
            "fellows": "Cohort_1_11_Demographics_latest.xlsx",
            "postdocs": "Postdocs_latest.xlsx",
            "grants": "Extra Grants_latest.xlsx",
            "institutionalization": "Institutionalization_latest.xlsx",
        },
        "rows": {
            "fellows": fellows,
            "postdocs": postdocs,
            "grants": grants,
            "trainings": trainings,
            "curricula_institutions": [{"name": n} for n in institutions],
        },
        "measures_static": {},
        "meta": {"summary": measures},
    }

    OUT_FILE.write_text(json.dumps(payload, ensure_ascii=False, default=str), encoding="utf-8")
    size_kb = OUT_FILE.stat().st_size / 1024
    print(f"Wrote {OUT_FILE} ({size_kb:.1f} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
